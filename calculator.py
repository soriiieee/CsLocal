import os,sys
import tkinter
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog

import openpyxl
import numpy as np # npという名前でimportする慣習。
from scipy import optimize
import copy

from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
cmap = plt.get_cmap("tab10")


from collections import defaultdict
import random

class Calculator:
    def __init__(self,excel_path=None,count_path=None):
        
        # self.excel_path = excel_path
        self.excel_path = os.path.abspath("CompileThemeStyle.xlsx")
        self.count_path = os.path.abspath("Titles.xlsx")
        
        # self.min_album = min_album
        self.music_dicts,self.colors,self.Remains_dicts = self._get_remains_dict()
        self.Quantity_dicts = self._get_themas_dict(self.music_dicts)
        self.album_titles = list(self.Quantity_dicts.keys())

    def _get_remains_dict(self):
        ## themas music nums ##
        Remains_dicts = {}
        df = pd.read_excel(self.excel_path,sheet_name="sheet2",skiprows=1)
        Remains_dicts = df.set_index("Style").to_dict()["Quantity"]
        
        i = 0
        colors,music_dicts = {},{}
        for thema in list(Remains_dicts.keys()):
            colors[thema] = cmap(i)
            music_dicts[thema] = i
            i += 1
        return music_dicts,colors,Remains_dicts

    def _get_themas_dict(self,music_dicts):
        ## themas music nums ##
        Quantity_dicts = {}
        df = pd.read_excel(self.excel_path,sheet_name="sheet1",skiprows=1)
        themas = df["FixTheme"].unique()
        for thema in themas:
            nums  = df[df["FixTheme"]==thema][["Style","Quantity"]].set_index("Style").to_dict()["Quantity"]
            for mc,count in music_dicts.items():
                if not mc in nums.keys():
                    nums[mc] = 0
            Quantity_dicts[thema] = nums
        return Quantity_dicts

    def plot_thema_pie(self,colors):
        N = len(list(self.Quantity_dicts.keys()))
        f,ax = plt.subplots(2,4,figsize=(22,11))
        ax = ax.flatten()
        for i,thema in enumerate(list(self.Quantity_dicts.keys())):
            label,x = self.Quantity_dicts[thema].keys(),self.Quantity_dicts[thema].values()
            label_mrow = [ x.replace("_","\n") for x in list(label)]
            ax[i].set_title(thema)
            ax[i].pie(x, labels=label_mrow, counterclock=False, startangle=90, 
                    shadow=False,colors=[ colors[t] for t in label ],
                    wedgeprops={'linewidth': 3, 'edgecolor':"white"},
                    )
        f.savefig("./sample.png" ,bbox_inches="tight")

    def get_music_df(self):
        # print(Quantity_dicts)
        df_nums = pd.DataFrame(self.Quantity_dicts).T
        return df_nums
    
    def get_needs(self):
        df = pd.read_excel(self.count_path,sheet_name="Sheet2")[["Title","Num2","use_album"]]
        gr = df.groupby("use_album").agg({"Num2":"sum"})
        gr["Num2"] = gr["Num2"].astype(int)
        return gr


    def calc_compile_number(self,values=None,available_songs2=None,titles_number2=None):   
         
        # plot..
        # colors = {  mc : cmap(i) for i,mc in enumerate(list(music_dicts.keys()))}
        # self.plot_thema_pie(colors)
        
        # quantity change
        # self.optimize_quantity(flags)
            
        # print(Quantity_dicts)
        df_nums = pd.DataFrame(self.Quantity_dicts).T
        self.df_nums = df_nums
        
        index = list(df_nums.index)
        columns = list(df_nums.columns)
        
        if values is not None:
            df_nums = pd.DataFrame(values,index=index , columns=columns)
            self.df_nums = df_nums
    
        
        ####
        # 線形計画法
        # https://qiita.com/NNNiNiNNN/items/57e409e5dbcfac9897ec
        # https://docs.scipy.org/doc/scipy/reference/generated/scipy.optimize.linprog.html

        C = np.ones(len(index)) * (-1)
        G = df_nums.values.T # 5 ,8
        
        h = np.array(list(self.Remains_dicts.values()))

        # アルバムの拘束条件
        bounds = tuple([ (0,need) for need in titles_number2])
        
        # exit()
        sol = optimize.linprog(C, G, h,bounds=bounds)

        albums = np.array([ int(np.floor(x)) for x in sol.x]) #float -> integer
        max_albums = np.sum(albums)
        
        # if mode == "test":
        #     albums = np.where(albums>min_album,min_album,min_album)
            
        use_musics = np.dot(G,albums)            
        remains_musics = available_songs2 - use_musics
        
        #output
        self.output_excel_sheet(use_musics,remains_musics,albums)
        print("Done...save excel")
        
        return albums ,use_musics, remains_musics
        

    
    def output_excel_sheet(self,use_musics,remains_musics,albums):
        
        ### output 
        wb = openpyxl.load_workbook(self.excel_path)
        if 'result' in wb.sheetnames:
            wb.remove(wb["result"])
        if 'quantities' in wb.sheetnames:
            wb.remove(wb["quantities"])
            
        ws = wb["sheet2"]
        ws["C2"],ws["D2"] = "UseNum" , f"Remains" 
        for i,(u,r) in enumerate(zip(use_musics,remains_musics)):
            ws[f"C{i+3}"],ws[f"D{i+3}"].value = u,r
            
        ws3 = wb.create_sheet("result")
        for c in ["A","B","C"]:
            ws3.column_dimensions[c].width = 20
        
        ws3["A1"],ws3["B1"] = "Albums" , f"MaxNum"    
        for i, (g,num) in enumerate(zip(self.album_titles,albums)):
            ws3[f"A{i+2}"],ws3[f"B{i+2}"] = g,num

        ws4 = wb.create_sheet("quantities")
        for c in ["A","B","C","D","E","F","G"]:
            ws4.column_dimensions[c].width = 25
        wb.save(self.excel_path)
        
        ##比率も保存する
        with pd.ExcelWriter(self.excel_path,mode='a',if_sheet_exists='overlay') as writer:
            self.df_nums.to_excel(writer, sheet_name= "quantities")

        
    def make_csv(self):
        """_summary_
        23.8.24 ysorimachi add orderの曲数を可視化するもの
        """
        create_date = datetime.now().strftime("%Y%m%d%H%M%S")
        df = pd.read_excel(self.excel_path,sheet_name="quantities")
        df = df.set_index("Unnamed: 0")
        df.index.name = "Title"
        df = df.T
                        
        titiles = []
        styles,quantities = [],[]
        
        choice_dict = defaultdict(lambda: defaultdict(int))
        order_sample_dict = defaultdict(lambda: defaultdict(str))
        for t in list(df.columns):
            dicts = df[t].to_dict()
            
            for (midi,count) in list(dicts.items()):        
                if count != 0:
                    titiles.append(t)
                    styles.append(midi)
                    quantities.append(count)
                    choice_dict[t][midi] = count
            
            
            best_score = 0
            ## order algorythm 
            for i in range(25):
                order , score = choice_random_order(choice_dict[t])
                if score > best_score:
                    best_score = score
                    best_order = order

            order_sample_dict[t] = best_order   
            
        # print(order_sample_dict)
        df_save = pd.DataFrame({"CompileTheme":titiles,"Style":styles,"Quantity":quantities})
                
        orders_str = [] 
        for i,r in df_save.iterrows():
            
            theme,style = r["CompileTheme"],r["Style"]
            ordr = order_sample_dict[theme][style]
            orders_str.append(ordr)
        
        df_save["Order"] = orders_str
        save_csv_name = f"./MusicEngine_CompileTheme-Style_{create_date}.csv"
        df_save.to_csv(save_csv_name,index=False)
        
        print("Done....")
        return df_save,save_csv_name
    

def choice_random_order(dicts={'Pf-Bs-Ds_BossaNova_4-4': 24, 'Pf-Bs-Perc_BossaNova_4-4': 6}):
    dicts2 = sorted(dicts.items(), key = lambda x: x[1])
    LIST = list(np.arange(1,30+1))
    
    dicts = {}
    N = len(dicts2)
    
    albums = {}
    cs = np.ones(31) * -1
    
    if N == 1:
        for ii,(style,num) in enumerate(dicts2):
            orders = sorted(random.sample(LIST,k=30))
            dicts[style] = ",".join(map(str , orders))
        return dicts,1
        
    for ii,(style,num) in enumerate(dicts2):
        if not style in albums:
            albums[style] = len(albums)
            
        if ii< N-1:
            trials = 0
            buffer = 3
            while True:
                orders = sorted(random.sample(LIST,k=num))
                
                if len(orders)>2:
                    flg = np.min(np.diff(orders))       
                else:
                    flg = 30         
                if flg > buffer and orders[0] !=1:
                    break
                
                trials +=1
                if trials>30:
                    buffer -=1
            LIST = sorted(list( set(LIST) - set(orders)))
        else:
            orders = sorted(random.sample(LIST,k=num))
        dicts[style] = ",".join(map(str , orders))
    
        for on in orders:
            cs[int(on)] = int(albums[style])
    
    score = np.abs(np.diff(cs[1:])).sum()
    return dicts,score
        
        
        

if __name__ == "__main__":
    choice_random_order()

    
    
    
